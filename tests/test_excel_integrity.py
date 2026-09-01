from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from excel_export import FIELD_WIDTHS, export_excel_outputs
from validate_excel_integrity import validate_against_canonical


def _raw_frame() -> pd.DataFrame:
    values = {
        "special": 7177,
        "prize1": 6825,
        "prize2_1": 8936,
        "prize2_2": 35676,
        "prize3_1": 1035,
        "prize3_2": 86863,
        "prize3_3": 36851,
        "prize3_4": 91550,
        "prize3_5": 59891,
        "prize3_6": 1824,
        "prize4_1": 946,
        "prize4_2": 9596,
        "prize4_3": 6005,
        "prize4_4": 872,
        "prize5_1": 23,
        "prize5_2": 3169,
        "prize5_3": 8526,
        "prize5_4": 486,
        "prize5_5": 7849,
        "prize5_6": 4836,
        "prize6_1": 82,
        "prize6_2": 824,
        "prize6_3": 787,
        "prize7_1": 2,
        "prize7_2": 92,
        "prize7_3": 40,
        "prize7_4": 0,
    }
    return pd.DataFrame([{"date": "2026-08-11", **values}])


def _two_frame(raw: pd.DataFrame) -> pd.DataFrame:
    two = raw.copy()
    for field in FIELD_WIDTHS:
        two[field] = two[field].astype(int) % 100
    return two


def _sparse_frame() -> pd.DataFrame:
    return pd.DataFrame([{"date": "2026-08-11", **{str(i): 0 for i in range(100)}}])


def test_excel_full_roundtrip_matches_canonical(tmp_path: Path) -> None:
    raw = _raw_frame()
    raw.to_csv(tmp_path / "xsmb.csv", index=False)
    export_excel_outputs(
        raw_df=raw,
        two_digit_df=_two_frame(raw),
        sparse_df=_sparse_frame(),
        data_dir=tmp_path,
        latest_daily_only=True,
    )
    report = validate_against_canonical(data_dir=tmp_path)
    assert report["ok"] is True
    assert report["canonical_rows"] == 1
    assert report["prize_values_per_draw"] == 27
    assert report["latest_daily_prize_values_checked"] == 27

    wb = load_workbook(tmp_path / "excel" / "xsmb.xlsx", data_only=False)
    ws = wb["Raw"]
    header = [c.value for c in ws[1]]
    for field, width in FIELD_WIDTHS.items():
        cell = ws.cell(2, header.index(field) + 1)
        assert isinstance(cell.value, str)
        assert len(cell.value) == width
        assert cell.value.isdigit()
        assert cell.number_format == "@"
    assert ws.cell(2, header.index("special") + 1).value == "07177"
    assert ws.cell(2, header.index("prize5_1") + 1).value == "0023"
    assert ws.cell(2, header.index("prize6_1") + 1).value == "082"
    assert ws.cell(2, header.index("prize7_1") + 1).value == "02"
    assert ws.cell(2, header.index("prize7_4") + 1).value == "00"


def test_excel_export_rejects_missing_prize_column(tmp_path: Path) -> None:
    raw = _raw_frame().drop(columns=["prize7_4"])
    with pytest.raises(ValueError, match="missing"):
        export_excel_outputs(
            raw_df=raw,
            two_digit_df=raw.copy(),
            sparse_df=_sparse_frame(),
            data_dir=tmp_path,
        )


def test_excel_export_rejects_dirty_or_overwidth_prize(tmp_path: Path) -> None:
    dirty = _raw_frame().astype(object)
    dirty.loc[0, "special"] = "07a17"
    with pytest.raises(ValueError, match="ASCII digits"):
        export_excel_outputs(
            raw_df=dirty,
            two_digit_df=_two_frame(_raw_frame()),
            sparse_df=_sparse_frame(),
            data_dir=tmp_path,
        )

    over = _raw_frame().astype(object)
    over.loc[0, "prize7_1"] = 100
    with pytest.raises(ValueError, match="outside width-2 range"):
        export_excel_outputs(
            raw_df=over,
            two_digit_df=_two_frame(_raw_frame()),
            sparse_df=_sparse_frame(),
            data_dir=tmp_path,
        )


def test_validator_detects_tampered_excel_value(tmp_path: Path) -> None:
    raw = _raw_frame()
    raw.to_csv(tmp_path / "xsmb.csv", index=False)
    export_excel_outputs(
        raw_df=raw,
        two_digit_df=_two_frame(raw),
        sparse_df=_sparse_frame(),
        data_dir=tmp_path,
        latest_daily_only=True,
    )
    path = tmp_path / "excel" / "xsmb.xlsx"
    wb = load_workbook(path)
    ws = wb["Raw"]
    header = [c.value for c in ws[1]]
    ws.cell(2, header.index("special") + 1).value = "07178"
    wb.save(path)
    with pytest.raises(RuntimeError, match="value mismatch"):
        validate_against_canonical(data_dir=tmp_path)
