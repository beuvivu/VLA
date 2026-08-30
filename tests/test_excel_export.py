from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path
from pathlib import Path as _Path

import pandas as pd

sys.path.insert(0, str((_Path(__file__).resolve().parents[1] / "src")))

from excel_export import export_excel_outputs


class TestExcelExport(unittest.TestCase):
    def test_excel_text_preserves_leading_zeroes(self):
        raw = pd.DataFrame(
            [
                {
                    "date": "2026-08-11",
                    "special": 7177,
                    "prize1": 6825,
                    "prize2_1": 8936,
                    "prize2_2": 35676,
                    "prize3_1": 1035,
                    "prize3_2": 86863,
                    "prize3_3": 36851,
                    "prize3_4": 91550,
                    "prize3_5": 59891,
                    "prize3_6": 1824,
                    "prize4_1": 946,
                    "prize4_2": 9596,
                    "prize4_3": 6005,
                    "prize4_4": 872,
                    "prize5_1": 23,
                    "prize5_2": 3169,
                    "prize5_3": 8526,
                    "prize5_4": 486,
                    "prize5_5": 7849,
                    "prize5_6": 4836,
                    "prize6_1": 82,
                    "prize6_2": 824,
                    "prize6_3": 787,
                    "prize7_1": 22,
                    "prize7_2": 92,
                    "prize7_3": 40,
                    "prize7_4": 0,
                }
            ]
        )
        two = raw.copy()
        for col in [c for c in two.columns if c != "date"]:
            two[col] = two[col] % 100
        sparse = pd.DataFrame([{"date": "2026-08-11", **{str(i): 0 for i in range(100)}}])
        with tempfile.TemporaryDirectory() as tmp:
            data_dir = Path(tmp)
            export_excel_outputs(raw_df=raw, two_digit_df=two, sparse_df=sparse, data_dir=data_dir, latest_daily_only=True)
            from openpyxl import load_workbook

            wb = load_workbook(data_dir / "excel" / "xsmb.xlsx", data_only=False)
            ws = wb["Raw"]
            header = [cell.value for cell in ws[1]]
            idx = header.index("special") + 1
            self.assertEqual(ws.cell(row=2, column=idx).value, "07177")
            self.assertEqual(ws.cell(row=2, column=idx).number_format, "@")


if __name__ == "__main__":
    unittest.main()
