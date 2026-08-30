from __future__ import annotations

"""Period matrices, boards, and AI/ML-assisted statistical signals for XSMB.

The tables in this module are descriptive analytics derived from historical
results. The ML-assisted scores are calibrated ranking signals, not guarantees.
"""

import argparse
import csv
import json
import logging
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Literal

import numpy as np
import pandas as pd

from lottery import Lottery, RepoPaths
from plot_utils import save_heatmap, save_labeled_heatmap, save_ranked_bar

PeriodKind = Literal["day", "week", "month", "year"]

logger = logging.getLogger(__name__)

WEEKDAY_COLS = ["T2", "T3", "T4", "T5", "T6", "T7", "CN"]
NUMBER_COLS = list(range(100))


@dataclass(frozen=True)
class StatsOutputs:
    data_dir: Path
    images_dir: Path
    excel_dir: Path

    @classmethod
    def from_paths(cls, paths: RepoPaths) -> "StatsOutputs":
        return cls(data_dir=paths.data_dir / "advanced", images_dir=paths.images_dir, excel_dir=paths.data_dir / "excel")


def _fmt2(n: int | float | str) -> str:
    return f"{int(n):02d}"


def _ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def _safe_float(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce").astype(float)


def _minmax(s: pd.Series) -> pd.Series:
    values = _safe_float(s)
    finite = values[np.isfinite(values)]
    if finite.empty:
        return pd.Series(np.zeros(len(values)), index=s.index)
    lo = float(finite.min())
    hi = float(finite.max())
    if np.isclose(lo, hi):
        return pd.Series(np.zeros(len(values)), index=s.index)
    return ((values - lo) / (hi - lo)).fillna(0.0)


def _prepare_sparse(sparse_df: pd.DataFrame) -> pd.DataFrame:
    """Normalize sparse columns to 0..99 integer labels."""
    if sparse_df.empty:
        return pd.DataFrame(columns=["date", *NUMBER_COLS])

    df = sparse_df.copy()
    df["date"] = pd.to_datetime(df["date"]).dt.normalize()

    rename: dict[object, int] = {}
    for col in df.columns:
        if col == "date":
            continue
        try:
            n = int(col)
        except Exception:
            continue
        if 0 <= n <= 99:
            rename[col] = n
    df = df.rename(columns=rename)

    for n in NUMBER_COLS:
        if n not in df.columns:
            df[n] = 0

    out = df[["date", *NUMBER_COLS]].copy()
    out[NUMBER_COLS] = out[NUMBER_COLS].apply(pd.to_numeric, errors="coerce").fillna(0).astype(int)
    return out.sort_values("date").reset_index(drop=True)


def _prepare_two_digits(two_digit_df: pd.DataFrame) -> pd.DataFrame:
    if two_digit_df.empty:
        return pd.DataFrame(columns=["date"])

    df = two_digit_df.copy()
    df["date"] = pd.to_datetime(df["date"]).dt.normalize()
    value_cols = [c for c in df.columns if c != "date"]
    df[value_cols] = df[value_cols].apply(pd.to_numeric, errors="coerce").fillna(0).astype(int) % 100
    return df.sort_values("date").reset_index(drop=True)


def _prepare_raw(raw_df: pd.DataFrame) -> pd.DataFrame:
    if raw_df.empty:
        return pd.DataFrame(columns=["date"])
    df = raw_df.copy()
    df["date"] = pd.to_datetime(df["date"]).dt.normalize()
    for c in df.columns:
        if c != "date":
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype(int)
    return df.sort_values("date").reset_index(drop=True)


def _de_sparse_from_raw(raw_df: pd.DataFrame) -> pd.DataFrame:
    df = _prepare_raw(raw_df)
    if df.empty or "special" not in df.columns:
        return pd.DataFrame(columns=["date", *NUMBER_COLS])
    arr = np.zeros((len(df), 100), dtype=np.int16)
    vals = (df["special"].astype(int).to_numpy() % 100).astype(int)
    arr[np.arange(len(vals)), vals] = 1
    return pd.concat([df[["date"]].reset_index(drop=True), pd.DataFrame(arr, columns=NUMBER_COLS)], axis=1)


def _first_prize_sparse_from_raw(raw_df: pd.DataFrame) -> pd.DataFrame:
    df = _prepare_raw(raw_df)
    if df.empty or "prize1" not in df.columns:
        return pd.DataFrame(columns=["date", *NUMBER_COLS])
    arr = np.zeros((len(df), 100), dtype=np.int16)
    vals = (df["prize1"].astype(int).to_numpy() % 100).astype(int)
    arr[np.arange(len(vals)), vals] = 1
    return pd.concat([df[["date"]].reset_index(drop=True), pd.DataFrame(arr, columns=NUMBER_COLS)], axis=1)


def _period_series(dates: pd.Series, period: PeriodKind) -> pd.Series:
    dt = pd.to_datetime(dates)
    if period == "day":
        return dt.dt.strftime("%Y-%m-%d")
    if period == "week":
        iso = dt.dt.isocalendar()
        return iso["year"].astype(str) + "-W" + iso["week"].astype(int).astype(str).str.zfill(2)
    if period == "month":
        return dt.dt.strftime("%Y-%m")
    if period == "year":
        return dt.dt.strftime("%Y")
    raise ValueError(f"Unsupported period: {period}")


def _period_frequency_from_sparse(sparse_df: pd.DataFrame, *, period: PeriodKind, mode: str) -> pd.DataFrame:
    df = _prepare_sparse(sparse_df)
    if df.empty:
        return pd.DataFrame()

    df["period_key"] = _period_series(df["date"], period)
    grouped_counts = df.groupby("period_key", sort=True)[NUMBER_COLS].sum()
    grouped_days = (df[NUMBER_COLS] > 0).assign(period_key=df["period_key"]).groupby("period_key", sort=True)[NUMBER_COLS].sum()
    draws = df.groupby("period_key", sort=True)["date"].nunique()

    count_long = grouped_counts.reset_index().melt(id_vars="period_key", var_name="number", value_name="freq")
    days_long = grouped_days.reset_index().melt(id_vars="period_key", var_name="number", value_name="days_hit")

    out = count_long.merge(days_long, on=["period_key", "number"], how="left")
    out["period_kind"] = period
    out["mode"] = mode
    out["draws"] = out["period_key"].map(draws.to_dict()).astype(int)
    out["number"] = out["number"].astype(int)
    out["number_str"] = out["number"].map(_fmt2)
    out["freq"] = out["freq"].astype(int)
    out["days_hit"] = out["days_hit"].astype(int)
    out["hit_rate"] = np.where(out["draws"] > 0, out["days_hit"] / out["draws"], 0.0)
    out["avg_per_draw"] = np.where(out["draws"] > 0, out["freq"] / out["draws"], 0.0)

    # Approximation: each draw has 27 two-digit positions for loto and one
    # position for de/prize1. This z-score is a comparison indicator only.
    positions_per_draw = 27 if mode == "loto" else 1
    n_positions = out["draws"].astype(float) * positions_per_draw
    expected = n_positions * 0.01
    denom = np.sqrt(np.maximum(n_positions * 0.01 * 0.99, 1e-9))
    out["expected_freq"] = expected.round(4)
    out["z_score"] = ((out["freq"] - expected) / denom).round(4)
    out["rank_in_period"] = out.groupby("period_key")["freq"].rank(method="dense", ascending=False).astype(int)

    cols = [
        "period_kind",
        "period_key",
        "mode",
        "draws",
        "number",
        "number_str",
        "freq",
        "days_hit",
        "hit_rate",
        "avg_per_draw",
        "expected_freq",
        "z_score",
        "rank_in_period",
    ]
    return out[cols].sort_values(["period_key", "rank_in_period", "number"]).reset_index(drop=True)


def _current_period_snapshot(freq_tables: dict[PeriodKind, pd.DataFrame], latest_date: pd.Timestamp, *, mode: str) -> pd.DataFrame:
    rows: list[pd.DataFrame] = []
    anchor = pd.Series([latest_date])
    for period, table in freq_tables.items():
        if table.empty:
            continue
        key = _period_series(anchor, period).iloc[0]
        cur = table[table["period_key"] == key].copy()
        cur["snapshot_as_of"] = latest_date.date().isoformat()
        cur["scope_label"] = {
            "day": "Ngày hiện tại",
            "week": "Tuần hiện tại",
            "month": "Tháng hiện tại",
            "year": "Năm hiện tại",
        }[period]
        cur["mode"] = mode
        rows.append(cur)
    if not rows:
        return pd.DataFrame()
    return pd.concat(rows, ignore_index=True).sort_values(["period_kind", "rank_in_period", "number"])


def _period_matrix(sparse_df: pd.DataFrame, *, period: PeriodKind, last_n: int | None = None) -> pd.DataFrame:
    df = _prepare_sparse(sparse_df)
    if df.empty:
        return pd.DataFrame()

    df["period_key"] = _period_series(df["date"], period)
    mat = df.groupby("period_key", sort=True)[NUMBER_COLS].sum().astype(int)
    mat.columns = [_fmt2(c) for c in mat.columns]
    if last_n is not None and len(mat) > last_n:
        mat = mat.tail(last_n)
    return mat


def _matrix_for_key(freq_table: pd.DataFrame, period_key: str) -> pd.DataFrame:
    cur = freq_table[freq_table["period_key"] == period_key].copy()
    if cur.empty:
        return pd.DataFrame()
    cur["tens"] = cur["number"] // 10
    cur["ones"] = cur["number"] % 10
    return cur.pivot(index="tens", columns="ones", values="freq").fillna(0).astype(int).sort_index()


def _rhythm_from_sparse(sparse_df: pd.DataFrame, *, mode: str) -> pd.DataFrame:
    df = _prepare_sparse(sparse_df)
    if df.empty:
        return pd.DataFrame()

    dates = pd.to_datetime(df["date"]).dt.normalize().to_numpy()
    as_of = pd.to_datetime(df["date"]).max().normalize()
    mat = df[NUMBER_COLS].to_numpy(dtype=int, copy=False)
    rows: list[dict[str, object]] = []

    for n in NUMBER_COLS:
        idx = np.where(mat[:, n] > 0)[0]
        if idx.size == 0:
            current_gap = int((as_of - pd.to_datetime(dates[0])).days)
            rows.append(
                {
                    "mode": mode,
                    "number": n,
                    "number_str": _fmt2(n),
                    "hit_count": 0,
                    "last_seen": None,
                    "current_gap": current_gap,
                    "min_gap": None,
                    "mean_gap": None,
                    "median_gap": None,
                    "max_gap": None,
                    "last_10_intervals": "",
                    "rhythm_pressure": None,
                }
            )
            continue

        hit_dates = pd.to_datetime(dates[idx]).normalize()
        gaps = np.diff(hit_dates).astype("timedelta64[D]").astype(int) if len(hit_dates) >= 2 else np.array([], dtype=int)
        current_gap = int((as_of - hit_dates[-1]).days)
        mean_gap = float(np.mean(gaps)) if gaps.size else None
        pressure = (current_gap / mean_gap) if mean_gap and mean_gap > 0 else None
        rows.append(
            {
                "mode": mode,
                "number": n,
                "number_str": _fmt2(n),
                "hit_count": int(idx.size),
                "last_seen": hit_dates[-1].date().isoformat(),
                "current_gap": current_gap,
                "min_gap": int(np.min(gaps)) if gaps.size else None,
                "mean_gap": None if mean_gap is None else round(mean_gap, 4),
                "median_gap": None if not gaps.size else round(float(np.median(gaps)), 4),
                "max_gap": int(np.max(gaps)) if gaps.size else None,
                "last_10_intervals": "|".join(str(int(x)) for x in gaps[-10:]),
                "rhythm_pressure": None if pressure is None else round(float(pressure), 4),
            }
        )

    out = pd.DataFrame(rows)
    return out.sort_values(["current_gap", "hit_count"], ascending=[False, True]).reset_index(drop=True)


def _special_boards(raw_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    raw = _prepare_raw(raw_df)
    if raw.empty or "special" not in raw.columns:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    df = raw[["date", "special"]].copy()
    df["special_2d"] = (df["special"].astype(int) % 100).map(_fmt2)
    df["special_full"] = df["special"].astype(int).map(lambda x: f"{x:05d}")
    df["week_key"] = _period_series(df["date"], "week")
    df["weekday"] = pd.to_datetime(df["date"]).dt.weekday.map({0: "T2", 1: "T3", 2: "T4", 3: "T5", 4: "T6", 5: "T7", 6: "CN"})
    df["month_key"] = _period_series(df["date"], "month")
    df["day"] = pd.to_datetime(df["date"]).dt.day.map(lambda d: f"{d:02d}")
    df["year"] = _period_series(df["date"], "year")

    week_board = (
        df.pivot_table(index="week_key", columns="weekday", values="special_2d", aggfunc="last")
        .reindex(columns=WEEKDAY_COLS)
        .sort_index()
        .reset_index()
    )

    day_cols = [f"{d:02d}" for d in range(1, 32)]
    month_board = (
        df.pivot_table(index="month_key", columns="day", values="special_2d", aggfunc="last")
        .reindex(columns=day_cols)
        .sort_index()
        .reset_index()
    )

    year_freq_rows: list[pd.DataFrame] = []
    month_freq_rows: list[pd.DataFrame] = []
    for period, key_col, rows in [("year", "year", year_freq_rows), ("month", "month_key", month_freq_rows)]:
        counts = df.groupby([key_col, "special_2d"], sort=True).size().reset_index(name="freq")
        counts["period_kind"] = period
        counts = counts.rename(columns={key_col: "period_key", "special_2d": "number_str"})
        counts["number"] = counts["number_str"].astype(int)
        counts["rank_in_period"] = counts.groupby("period_key")["freq"].rank(method="dense", ascending=False).astype(int)
        rows.append(counts[["period_kind", "period_key", "number", "number_str", "freq", "rank_in_period"]])
    return week_board, month_board, year_freq_rows[0], month_freq_rows[0]


def _special_group_frequency(raw_df: pd.DataFrame, *, period: PeriodKind) -> pd.DataFrame:
    raw = _prepare_raw(raw_df)
    if raw.empty or "special" not in raw.columns:
        return pd.DataFrame()

    df = raw[["date", "special"]].copy()
    df["period_key"] = _period_series(df["date"], period)
    last2 = (df["special"].astype(int) % 100).astype(int)
    tens = last2 // 10
    ones = last2 % 10

    frames = [
        pd.DataFrame({"period_key": df["period_key"], "group_type": "db_head", "group_value": tens}),
        pd.DataFrame({"period_key": df["period_key"], "group_type": "db_tail", "group_value": ones}),
        pd.DataFrame({"period_key": df["period_key"], "group_type": "db_total", "group_value": (tens + ones) % 10}),
    ]

    cham_rows: list[dict[str, object]] = []
    for key, a, b in zip(df["period_key"], tens, ones, strict=False):
        for digit in sorted({int(a), int(b)}):
            cham_rows.append({"period_key": key, "group_type": "db_cham", "group_value": digit})
    frames.append(pd.DataFrame(cham_rows))

    long = pd.concat(frames, ignore_index=True)
    counts = long.groupby(["period_key", "group_type", "group_value"], sort=True).size().reset_index(name="freq")
    draws = df.groupby("period_key")["date"].nunique()
    counts["period_kind"] = period
    counts["draws"] = counts["period_key"].map(draws.to_dict()).astype(int)
    counts["rate"] = np.where(counts["draws"] > 0, counts["freq"] / counts["draws"], 0.0)
    counts["rank_in_period_group"] = counts.groupby(["period_key", "group_type"])["freq"].rank(method="dense", ascending=False).astype(int)
    return counts[["period_kind", "period_key", "group_type", "group_value", "freq", "draws", "rate", "rank_in_period_group"]]


def _head_tail_total_by_period(two_digit_df: pd.DataFrame, *, period: PeriodKind) -> pd.DataFrame:
    two = _prepare_two_digits(two_digit_df)
    if two.empty:
        return pd.DataFrame()

    value_cols = [c for c in two.columns if c != "date"]
    long = two[["date", *value_cols]].melt(id_vars="date", value_name="number")
    long["number"] = long["number"].astype(int) % 100
    long["period_key"] = _period_series(long["date"], period)
    long["head"] = long["number"] // 10
    long["tail"] = long["number"] % 10
    long["total"] = (long["head"] + long["tail"]) % 10

    rows: list[pd.DataFrame] = []
    for group_type, col in [("head", "head"), ("tail", "tail"), ("total", "total")]:
        counts = long.groupby(["period_key", col], sort=True).size().reset_index(name="freq")
        counts["period_kind"] = period
        counts["group_type"] = group_type
        counts = counts.rename(columns={col: "group_value"})
        counts["rank_in_period_group"] = counts.groupby("period_key")["freq"].rank(method="dense", ascending=False).astype(int)
        rows.append(counts[["period_kind", "period_key", "group_type", "group_value", "freq", "rank_in_period_group"]])
    return pd.concat(rows, ignore_index=True).sort_values(["period_key", "group_type", "rank_in_period_group"])


def _reverse_pair_frequency(sparse_df: pd.DataFrame, *, period: PeriodKind) -> pd.DataFrame:
    df = _prepare_sparse(sparse_df)
    if df.empty:
        return pd.DataFrame()

    df["period_key"] = _period_series(df["date"], period)
    count_by_period = df.groupby("period_key", sort=True)[NUMBER_COLS].sum()
    bool_df = (df[NUMBER_COLS] > 0).assign(period_key=df["period_key"])
    days_by_period = bool_df.groupby("period_key", sort=True)[NUMBER_COLS].sum()
    draws = df.groupby("period_key", sort=True)["date"].nunique()

    pairs: list[tuple[int, int]] = []
    seen: set[tuple[int, int]] = set()
    for n in NUMBER_COLS:
        rev = (n % 10) * 10 + (n // 10)
        key = tuple(sorted((n, rev)))
        if key not in seen:
            seen.add(key)
            pairs.append(key)

    rows: list[dict[str, object]] = []
    for period_key in count_by_period.index:
        for a, b in pairs:
            if a == b:
                freq = int(count_by_period.loc[period_key, a])
                days_hit = int(days_by_period.loc[period_key, a])
                cooccur_days = days_hit
            else:
                freq = int(count_by_period.loc[period_key, a] + count_by_period.loc[period_key, b])
                a_hit = df.loc[df["period_key"] == period_key, a] > 0
                b_hit = df.loc[df["period_key"] == period_key, b] > 0
                days_hit = int((a_hit | b_hit).sum())
                cooccur_days = int((a_hit & b_hit).sum())
            rows.append(
                {
                    "period_kind": period,
                    "period_key": period_key,
                    "draws": int(draws.loc[period_key]),
                    "pair": f"{a:02d}-{b:02d}",
                    "a": a,
                    "b": b,
                    "freq": freq,
                    "days_hit": days_hit,
                    "cooccur_days": cooccur_days,
                    "avg_per_draw": freq / max(int(draws.loc[period_key]), 1),
                }
            )

    out = pd.DataFrame(rows)
    out["rank_in_period"] = out.groupby("period_key")["freq"].rank(method="dense", ascending=False).astype(int)
    return out.sort_values(["period_key", "rank_in_period", "pair"]).reset_index(drop=True)


def _first_prize_stats(raw_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    sparse = _first_prize_sparse_from_raw(raw_df)
    freq = _period_frequency_from_sparse(sparse, period="year", mode="prize1")
    rhythm = _rhythm_from_sparse(sparse, mode="prize1")
    overdue = rhythm[["mode", "number", "number_str", "last_seen", "current_gap", "hit_count", "mean_gap", "max_gap"]].copy()
    return freq, overdue.sort_values("current_gap", ascending=False)


def _conditional_tables(raw_df: pd.DataFrame, sparse_df: pd.DataFrame, *, top: int) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    raw = _prepare_raw(raw_df)
    sparse = _prepare_sparse(sparse_df)
    if raw.empty or sparse.empty or len(raw) < 2:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    raw = raw.reset_index(drop=True)
    sparse = sparse.reset_index(drop=True)
    prev_de = (raw["special"].astype(int).to_numpy() % 100).astype(int)
    next_de = prev_de[1:]
    mat = (sparse[NUMBER_COLS].to_numpy(dtype=int, copy=False) > 0).astype(np.int16)

    # next-day loto conditional on previous special/de
    de_loto = np.zeros((100, 100), dtype=np.int32)
    de_de = np.zeros((100, 100), dtype=np.int32)
    prev_loto_next_loto = np.zeros((100, 100), dtype=np.int32)
    prev_de_counts = np.zeros(100, dtype=np.int32)
    prev_loto_counts = np.zeros(100, dtype=np.int32)

    for t in range(len(raw) - 1):
        d = prev_de[t]
        prev_de_counts[d] += 1
        next_hits = np.where(mat[t + 1] > 0)[0]
        de_loto[d, next_hits] += 1
        de_de[d, next_de[t]] += 1

        prev_hits = np.where(mat[t] > 0)[0]
        prev_loto_counts[prev_hits] += 1
        if len(prev_hits) and len(next_hits):
            prev_loto_next_loto[np.ix_(prev_hits, next_hits)] += 1

    def top_matrix(mat_counts: np.ndarray, row_counts: np.ndarray, row_name: str, col_name: str) -> pd.DataFrame:
        rows: list[dict[str, object]] = []
        for i in range(100):
            denom = int(row_counts[i])
            if denom == 0:
                continue
            for j in range(100):
                c = int(mat_counts[i, j])
                if c <= 0:
                    continue
                rows.append(
                    {
                        row_name: _fmt2(i),
                        col_name: _fmt2(j),
                        "count": c,
                        "base_count": denom,
                        "conditional_rate": c / denom,
                    }
                )
        return pd.DataFrame(rows).sort_values(["conditional_rate", "count"], ascending=[False, False]).head(top).reset_index(drop=True)

    return (
        top_matrix(de_loto, prev_de_counts, "prev_special_2d", "next_loto"),
        top_matrix(de_de, prev_de_counts, "prev_special_2d", "next_special_2d"),
        top_matrix(prev_loto_next_loto, prev_loto_counts, "prev_loto", "next_loto"),
    )


def _load_ml_probs(data_dir: Path, mode: str) -> pd.DataFrame:
    candidates = [
        # Prefer the newest explainable cầu-kèo AI/ML layer when available.
        data_dir / "ai_ml" / f"cau_keo_{mode}_all.csv",
        data_dir / "ml" / f"predict_next_{mode}_ml_all.csv",
        data_dir / "predict" / f"predict_next_{mode}_all.csv",
    ]
    # Also support date-stamped ensemble predictions.
    candidates.extend(sorted((data_dir / "predict").glob(f"predict_next_{mode}_all_*.csv"), reverse=True))

    for path in candidates:
        if not path.exists():
            continue
        try:
            df = pd.read_csv(path, dtype={"number": str})
            if "number" not in df.columns or "prob" not in df.columns:
                continue
            out = df[["number", "prob"]].copy()
            out["number_str"] = out["number"].astype(str).str.extract(r"(\d+)")[0].fillna("0").astype(int).map(_fmt2)
            out["ml_prob"] = pd.to_numeric(out["prob"], errors="coerce").fillna(0.0)
            return out[["number_str", "ml_prob"]].groupby("number_str", as_index=False)["ml_prob"].max()
        except Exception as exc:  # noqa: BLE001
            logger.debug("Cannot load ML probabilities from %s: %s", path, exc)
    return pd.DataFrame({"number_str": [_fmt2(i) for i in NUMBER_COLS], "ml_prob": np.zeros(100)})


def _window_freq(sparse_df: pd.DataFrame, days: int, *, mode: str) -> pd.DataFrame:
    df = _prepare_sparse(sparse_df)
    if df.empty:
        return pd.DataFrame()
    max_date = df["date"].max()
    window = df[(df["date"] > max_date - pd.Timedelta(days=days)) & (df["date"] <= max_date)]
    counts = window[NUMBER_COLS].sum(axis=0).astype(int)
    days_hit = (window[NUMBER_COLS] > 0).sum(axis=0).astype(int)
    out = pd.DataFrame(
        {
            "mode": mode,
            "number": NUMBER_COLS,
            "number_str": [_fmt2(i) for i in NUMBER_COLS],
            f"freq_{days}d": counts.to_numpy(),
            f"days_hit_{days}d": days_hit.to_numpy(),
        }
    )
    return out


def _ai_ml_signal(
    *,
    data_dir: Path,
    mode: str,
    sparse_df: pd.DataFrame,
    rhythm_df: pd.DataFrame,
    year_freq_table: pd.DataFrame,
) -> pd.DataFrame:
    ml = _load_ml_probs(data_dir, mode)
    f7 = _window_freq(sparse_df, 7, mode=mode)
    f30 = _window_freq(sparse_df, 30, mode=mode)
    f365 = _window_freq(sparse_df, 365, mode=mode)

    out = pd.DataFrame({"number": NUMBER_COLS, "number_str": [_fmt2(i) for i in NUMBER_COLS]})
    for part in [f7, f30, f365]:
        cols = [c for c in part.columns if c not in {"mode", "number", "number_str"}]
        out = out.merge(part[["number", *cols]], on="number", how="left")
    out = out.merge(rhythm_df[["number", "current_gap", "mean_gap", "rhythm_pressure"]], on="number", how="left")
    out = out.merge(ml, on="number_str", how="left")

    # Year frequency from current/latest year for prior strength.
    if not year_freq_table.empty:
        latest_year = sorted(year_freq_table["period_key"].astype(str).unique())[-1]
        yf = year_freq_table[year_freq_table["period_key"].astype(str) == latest_year][["number", "freq", "z_score"]].copy()
        yf = yf.rename(columns={"freq": "freq_current_year", "z_score": "z_score_current_year"})
        out = out.merge(yf, on="number", how="left")
    else:
        out["freq_current_year"] = 0
        out["z_score_current_year"] = 0.0

    for c in ["freq_7d", "freq_30d", "freq_365d", "current_gap", "mean_gap", "rhythm_pressure", "ml_prob", "freq_current_year", "z_score_current_year"]:
        if c not in out.columns:
            out[c] = 0
        out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)

    ml_norm = _minmax(out["ml_prob"])
    freq30_norm = _minmax(out["freq_30d"])
    freq7_norm = _minmax(out["freq_7d"])
    year_norm = _minmax(out["freq_current_year"])
    pressure_norm = _minmax(out["rhythm_pressure"].clip(lower=0, upper=5))
    gap_norm = _minmax(out["current_gap"])

    if mode == "loto":
        score = 100 * (0.40 * ml_norm + 0.25 * freq30_norm + 0.20 * pressure_norm + 0.15 * freq7_norm)
    else:
        score = 100 * (0.45 * ml_norm + 0.20 * year_norm + 0.25 * gap_norm + 0.10 * freq30_norm)

    out["ai_ml_signal_score"] = score.round(4)
    out["ml_prob"] = out["ml_prob"].round(8)
    out["score_band"] = pd.cut(
        out["ai_ml_signal_score"],
        bins=[-0.01, 25, 50, 75, 100.01],
        labels=["low", "medium", "high", "very_high"],
    ).astype(str)
    out["note"] = "statistical ranking signal only; not a guaranteed prediction"
    return out.sort_values(["ai_ml_signal_score", "ml_prob"], ascending=False).reset_index(drop=True)


def _to_10x10_from_number_metric(df: pd.DataFrame, *, metric: str) -> pd.DataFrame:
    tmp = df.copy()
    tmp["number"] = tmp["number"].astype(int)
    tmp["tens"] = tmp["number"] // 10
    tmp["ones"] = tmp["number"] % 10
    return tmp.pivot(index="tens", columns="ones", values=metric).fillna(0).astype(float).sort_index()


def _dump_table(df: pd.DataFrame, *, out_dir: Path, name: str) -> Path:
    _ensure_dir(out_dir)
    path = out_dir / f"{name}.csv"
    df.to_csv(path, index=False, quoting=csv.QUOTE_NONNUMERIC)
    df.to_json(out_dir / f"{name}.json", orient="records", indent=2, force_ascii=False)
    return path


def _write_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as exc:  # pragma: no cover
        logger.warning("openpyxl not available; skip statistics workbook: %s", exc)
        return

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="172554")
    header_font = Font(color="FFFFFF", bold=True)
    text_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])
        if df.empty:
            ws.cell(row=1, column=1, value="No data")
            continue

        for col_idx, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=str(col))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = text_align

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                col_name = str(df.columns[col_idx - 1])
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if isinstance(value, str) or col_name.endswith("_str") or col_name in {"number", "pair"}:
                    cell.number_format = "@"
                    cell.alignment = text_align
                elif isinstance(value, (int, float, np.integer, np.floating)):
                    cell.alignment = right_align
                else:
                    cell.alignment = text_align

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col_idx, col in enumerate(df.columns, start=1):
            values = df.iloc[:500, col_idx - 1].astype(str).tolist()
            width = min(max(10, max([len(str(col)), *[len(v) for v in values]]) + 2), 28)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(path)


def _manifest(*, out_dir: Path, as_of: str, files: list[str]) -> None:
    payload = {
        "generated_at": datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "as_of_date": as_of,
        "files": sorted(files),
        "note": "All outputs are historical/statistical analytics. ML scores are ranking signals only and do not guarantee future lottery outcomes.",
    }
    (out_dir / "statistics_manifest.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def generate_statistics(
    *,
    freq_days: int = 99,
    top_conditional: int = 500,
    recent_days: int = 31,
    recent_weeks: int = 26,
    recent_months: int = 24,
) -> list[Path]:
    paths = RepoPaths.from_module()
    out = StatsOutputs.from_paths(paths)
    _ensure_dir(out.data_dir)
    _ensure_dir(out.images_dir)

    lot = Lottery()
    lot.load()
    raw = _prepare_raw(lot.get_raw_data())
    two = _prepare_two_digits(lot.get_2_digits_data())
    sparse_loto = _prepare_sparse(lot.get_sparse_data())
    sparse_de = _de_sparse_from_raw(raw)

    if raw.empty or sparse_loto.empty:
        raise RuntimeError("No data loaded. Run src/sync.py first.")

    latest_date = pd.to_datetime(raw["date"]).max().normalize()
    latest_iso = latest_date.date().isoformat()
    created: list[Path] = []

    # 1) Period frequency tables and matrices: loto + de.
    freq_loto: dict[PeriodKind, pd.DataFrame] = {}
    freq_de: dict[PeriodKind, pd.DataFrame] = {}
    for period in ["day", "week", "month", "year"]:
        p = period  # type: ignore[assignment]
        freq_loto[p] = _period_frequency_from_sparse(sparse_loto, period=p, mode="loto")
        freq_de[p] = _period_frequency_from_sparse(sparse_de, period=p, mode="de")
        created.append(_dump_table(freq_loto[p], out_dir=out.data_dir, name=f"period_frequency_loto_{period}"))
        created.append(_dump_table(freq_de[p], out_dir=out.data_dir, name=f"period_frequency_de_{period}"))

    snapshot_loto = _current_period_snapshot(freq_loto, latest_date, mode="loto")
    snapshot_de = _current_period_snapshot(freq_de, latest_date, mode="de")
    created.append(_dump_table(snapshot_loto, out_dir=out.data_dir, name="period_snapshot_loto_current"))
    created.append(_dump_table(snapshot_de, out_dir=out.data_dir, name="period_snapshot_de_current"))

    matrix_specs = [
        ("day", recent_days, f"last{recent_days}"),
        ("week", recent_weeks, f"last{recent_weeks}"),
        ("month", recent_months, f"last{recent_months}"),
        ("year", None, "all"),
    ]
    for period, last_n, suffix in matrix_specs:
        p = period  # type: ignore[assignment]
        m_loto = _period_matrix(sparse_loto, period=p, last_n=last_n)
        m_de = _period_matrix(sparse_de, period=p, last_n=last_n)
        created.append(_dump_table(m_loto.reset_index(), out_dir=out.data_dir, name=f"period_matrix_loto_{period}_{suffix}"))
        created.append(_dump_table(m_de.reset_index(), out_dir=out.data_dir, name=f"period_matrix_de_{period}_{suffix}"))

    # 2) Timeframe 10x10 heatmaps for the current day/week/month/year.
    for period, label in [
        ("day", "ngày mới nhất"),
        ("week", "tuần hiện tại"),
        ("month", "tháng hiện tại"),
        ("year", "năm hiện tại"),
    ]:
        p = period  # type: ignore[assignment]
        key = _period_series(pd.Series([latest_date]), p).iloc[0]
        heat = _matrix_for_key(freq_loto[p], key)
        if not heat.empty:
            img = out.images_dir / f"period_loto_current_{period}_heatmap.jpg"
            save_heatmap(heat, title=f"Tần suất loto — {label}", out_path=img, cmap="YlGnBu", cbar_label="Số lần")
        heat_de = _matrix_for_key(freq_de[p], key)
        if not heat_de.empty and period in {"month", "year"}:
            img = out.images_dir / f"period_de_current_{period}_heatmap.jpg"
            save_heatmap(heat_de, title=f"Tần suất ĐB 2 số — {label}", out_path=img, cmap="YlOrRd", cbar_label="Số lần")

    weekly_mat = _period_matrix(sparse_loto, period="week", last_n=min(recent_weeks, 16))
    if not weekly_mat.empty:
        top_numbers = snapshot_loto[snapshot_loto["period_kind"] == "year"].sort_values("freq", ascending=False)["number_str"].head(20)
        trend = weekly_mat[[c for c in top_numbers if c in weekly_mat.columns]]
        save_labeled_heatmap(
            trend,
            title="Ma trận xu hướng tuần — top loto trong năm",
            out_path=out.images_dir / "period_loto_weekly_trend_top20.jpg",
            cmap="YlGnBu",
            cbar_label="Số lần/tuần",
            xlabel="Bộ số",
            ylabel="Tuần",
            annotate=False,
        )

    # 3) Rhythm/gan, head-tail-total, cặp lộn.
    rhythm_loto = _rhythm_from_sparse(sparse_loto, mode="loto")
    rhythm_de = _rhythm_from_sparse(sparse_de, mode="de")
    created.append(_dump_table(rhythm_loto, out_dir=out.data_dir, name="loto_rhythm"))
    created.append(_dump_table(rhythm_de, out_dir=out.data_dir, name="de_rhythm"))

    save_heatmap(
        _to_10x10_from_number_metric(rhythm_loto, metric="current_gap"),
        title="Nhịp loto / lô gan — số ngày chưa về",
        out_path=out.images_dir / "loto_rhythm_current_gap_heatmap.jpg",
        cmap="YlOrRd",
        cbar_label="Số ngày",
    )
    save_heatmap(
        _to_10x10_from_number_metric(rhythm_de, metric="current_gap"),
        title="Nhịp ĐB — số ngày chưa về",
        out_path=out.images_dir / "de_rhythm_current_gap_heatmap.jpg",
        cmap="YlOrRd",
        cbar_label="Số ngày",
    )

    hht_frames: list[pd.DataFrame] = []
    pair_frames: list[pd.DataFrame] = []
    special_group_frames: list[pd.DataFrame] = []
    for period in ["day", "week", "month", "year"]:
        p = period  # type: ignore[assignment]
        hht = _head_tail_total_by_period(two, period=p)
        pair = _reverse_pair_frequency(sparse_loto, period=p)
        sg = _special_group_frequency(raw, period=p)
        created.append(_dump_table(hht, out_dir=out.data_dir, name=f"head_tail_total_loto_{period}"))
        created.append(_dump_table(pair, out_dir=out.data_dir, name=f"reverse_pair_frequency_{period}"))
        created.append(_dump_table(sg, out_dir=out.data_dir, name=f"special_group_frequency_{period}"))
        hht_frames.append(hht)
        pair_frames.append(pair)
        special_group_frames.append(sg)

    current_keys = {p: _period_series(pd.Series([latest_date]), p).iloc[0] for p in ["day", "week", "month", "year"]}
    hht_current = pd.concat(
        [df[df["period_key"] == current_keys[df["period_kind"].iloc[0]]] for df in hht_frames if not df.empty],
        ignore_index=True,
    )
    pair_current = pd.concat(
        [df[df["period_key"] == current_keys[df["period_kind"].iloc[0]]] for df in pair_frames if not df.empty],
        ignore_index=True,
    )
    sg_current = pd.concat(
        [df[df["period_key"] == current_keys[df["period_kind"].iloc[0]]] for df in special_group_frames if not df.empty],
        ignore_index=True,
    )
    created.append(_dump_table(hht_current, out_dir=out.data_dir, name="head_tail_total_loto_current"))
    created.append(_dump_table(pair_current, out_dir=out.data_dir, name="reverse_pair_frequency_current"))
    created.append(_dump_table(sg_current, out_dir=out.data_dir, name="special_group_frequency_current"))

    if not pair_current.empty:
        pair_month = pair_current[pair_current["period_kind"] == "month"]
        if not pair_month.empty:
            save_ranked_bar(
                pair_month,
                x="pair",
                y="freq",
                title="Top cặp lộn theo tần suất — tháng hiện tại",
                out_path=out.images_dir / "reverse_pair_current_month_top.jpg",
                top_n=20,
                cmap="YlGnBu",
                ylabel="Số lần",
            )

    # 4) Special prize boards and first prize.
    week_board, month_board, special_year_freq, special_month_freq = _special_boards(raw)
    created.append(_dump_table(week_board, out_dir=out.data_dir, name="special_week_board"))
    created.append(_dump_table(month_board, out_dir=out.data_dir, name="special_month_board"))
    created.append(_dump_table(special_year_freq, out_dir=out.data_dir, name="special_year_frequency"))
    created.append(_dump_table(special_month_freq, out_dir=out.data_dir, name="special_month_frequency"))

    first_freq, first_overdue = _first_prize_stats(raw)
    created.append(_dump_table(first_freq, out_dir=out.data_dir, name="first_prize_year_frequency"))
    created.append(_dump_table(first_overdue, out_dir=out.data_dir, name="first_prize_overdue"))

    # 5) Conditional tables: loto theo đặc biệt, đề theo đặc biệt, loto theo loto.
    cond_de_loto, cond_de_de, cond_loto_loto = _conditional_tables(raw, sparse_loto, top=top_conditional)
    created.append(_dump_table(cond_de_loto, out_dir=out.data_dir, name="conditional_loto_after_special_top500"))
    created.append(_dump_table(cond_de_de, out_dir=out.data_dir, name="conditional_special_after_special_top500"))
    created.append(_dump_table(cond_loto_loto, out_dir=out.data_dir, name="conditional_loto_after_loto_top500"))

    # 6) AI/ML signal overlays. Uses ML output when present; otherwise zeros.
    signal_loto = _ai_ml_signal(
        data_dir=paths.data_dir,
        mode="loto",
        sparse_df=sparse_loto,
        rhythm_df=rhythm_loto,
        year_freq_table=freq_loto["year"],
    )
    signal_de = _ai_ml_signal(
        data_dir=paths.data_dir,
        mode="de",
        sparse_df=sparse_de,
        rhythm_df=rhythm_de,
        year_freq_table=freq_de["year"],
    )
    created.append(_dump_table(signal_loto, out_dir=out.data_dir, name="ai_ml_signal_loto"))
    created.append(_dump_table(signal_de, out_dir=out.data_dir, name="ai_ml_signal_de"))

    save_heatmap(
        _to_10x10_from_number_metric(signal_loto, metric="ai_ml_signal_score"),
        title="AI/ML signal loto — score tổng hợp",
        out_path=out.images_dir / "ai_ml_signal_loto_heatmap.jpg",
        cmap="YlGnBu",
        cbar_label="Signal score",
    )
    save_heatmap(
        _to_10x10_from_number_metric(signal_de, metric="ai_ml_signal_score"),
        title="AI/ML signal ĐB — score tổng hợp",
        out_path=out.images_dir / "ai_ml_signal_de_heatmap.jpg",
        cmap="YlOrRd",
        cbar_label="Signal score",
    )

    # 7) Excel-safe workbook for analysts.
    _write_excel(
        out.excel_dir / "statistics_matrices.xlsx",
        {
            "Snapshot Loto": snapshot_loto.head(400),
            "Snapshot De": snapshot_de.head(400),
            "AI Loto": signal_loto.head(100),
            "AI De": signal_de.head(100),
            "Loto Rhythm": rhythm_loto.head(100),
            "De Rhythm": rhythm_de.head(100),
            "Pair Current": pair_current.head(250),
            "HeadTail Current": hht_current.head(250),
            "DB Week Board": week_board.tail(40),
            "DB Month Board": month_board.tail(24),
            "First Overdue": first_overdue.head(100),
        },
    )

    _manifest(out_dir=out.data_dir, as_of=latest_iso, files=[str(p.relative_to(paths.root)) for p in created])
    logger.info("Generated statistical matrices as of %s under %s", latest_iso, out.data_dir)
    return created


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate daily/weekly/monthly/yearly statistical matrices for XSMB.")
    parser.add_argument("--freq-days", type=int, default=99, help="Compatibility option; retained for scheduling.")
    parser.add_argument("--top-conditional", type=int, default=500)
    parser.add_argument("--recent-days", type=int, default=31)
    parser.add_argument("--recent-weeks", type=int, default=26)
    parser.add_argument("--recent-months", type=int, default=24)
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    generate_statistics(
        freq_days=args.freq_days,
        top_conditional=args.top_conditional,
        recent_days=args.recent_days,
        recent_weeks=args.recent_weeks,
        recent_months=args.recent_months,
    )


if __name__ == "__main__":
    main()
