from __future__ import annotations

import logging
import math
from numbers import Integral, Real
from pathlib import Path

import pandas as pd

logger = logging.getLogger(__name__)

FIELD_WIDTHS: dict[str, int] = {
    "special": 5,
    "prize1": 5,
    "prize2_1": 5,
    "prize2_2": 5,
    "prize3_1": 5,
    "prize3_2": 5,
    "prize3_3": 5,
    "prize3_4": 5,
    "prize3_5": 5,
    "prize3_6": 5,
    "prize4_1": 4,
    "prize4_2": 4,
    "prize4_3": 4,
    "prize4_4": 4,
    "prize5_1": 4,
    "prize5_2": 4,
    "prize5_3": 4,
    "prize5_4": 4,
    "prize5_5": 4,
    "prize5_6": 4,
    "prize6_1": 3,
    "prize6_2": 3,
    "prize6_3": 3,
    "prize7_1": 2,
    "prize7_2": 2,
    "prize7_3": 2,
    "prize7_4": 2,
}

PRIZE_GROUPS: list[tuple[str, list[str]]] = [
    ("Đặc biệt", ["special"]),
    ("Giải nhất", ["prize1"]),
    ("Giải nhì", ["prize2_1", "prize2_2"]),
    (
        "Giải ba",
        [
            "prize3_1",
            "prize3_2",
            "prize3_3",
            "prize3_4",
            "prize3_5",
            "prize3_6",
        ],
    ),
    ("Giải tư", ["prize4_1", "prize4_2", "prize4_3", "prize4_4"]),
    (
        "Giải năm",
        [
            "prize5_1",
            "prize5_2",
            "prize5_3",
            "prize5_4",
            "prize5_5",
            "prize5_6",
        ],
    ),
    ("Giải sáu", ["prize6_1", "prize6_2", "prize6_3"]),
    ("Giải bảy", ["prize7_1", "prize7_2", "prize7_3", "prize7_4"]),
]

EXPECTED_PRIZE_VALUES = sum(len(fields) for _, fields in PRIZE_GROUPS)
if EXPECTED_PRIZE_VALUES != 27:
    raise RuntimeError("XSMB prize schema must contain exactly 27 values")


def _fmt_date(v: object) -> str:
    return pd.to_datetime(v).date().isoformat()


def _integer_value(value: object, *, label: str) -> int:
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        raise ValueError(f"{label} cannot be null")
    if isinstance(value, bool):
        raise ValueError(f"{label} cannot be boolean")
    if isinstance(value, str):
        raw = value.strip()
        if not raw or not raw.isascii() or not raw.isdigit():
            raise ValueError(f"{label} must contain ASCII digits only: {value!r}")
        return int(raw)
    if isinstance(value, Integral):
        return int(value)
    if isinstance(value, Real):
        f = float(value)
        if not math.isfinite(f) or not f.is_integer():
            raise ValueError(f"{label} must be an integer value: {value!r}")
        return int(f)
    raise ValueError(f"{label} has unsupported type: {type(value).__name__}")


def _fmt_prize(value: object, field: str) -> str:
    width = FIELD_WIDTHS[field]
    n = _integer_value(value, label=field)
    if not 0 <= n < 10**width:
        raise ValueError(
            f"{field} outside width-{width} range 0..{10**width - 1}: {value!r}"
        )
    return f"{n:0{width}d}"


def _fmt_two_digit(value: object, field: str) -> str:
    n = _integer_value(value, label=field)
    if not 0 <= n <= 99:
        raise ValueError(f"{field} two-digit value outside 00..99: {value!r}")
    return f"{n:02d}"


def _require_prize_columns(df: pd.DataFrame, *, context: str) -> None:
    missing = [field for field in FIELD_WIDTHS if field not in df.columns]
    if missing:
        raise ValueError(
            f"{context} is missing {len(missing)} prize column(s): {missing}"
        )


def _format_raw_df(df: pd.DataFrame) -> pd.DataFrame:
    _require_prize_columns(df, context="raw dataframe")
    out = df.copy()
    if "date" not in out.columns:
        raise ValueError("raw dataframe is missing date")
    out["date"] = out["date"].apply(_fmt_date)
    for col in FIELD_WIDTHS:
        out[col] = out[col].apply(lambda v, field=col: _fmt_prize(v, field))
    return out


def _format_two_digit_df(df: pd.DataFrame) -> pd.DataFrame:
    _require_prize_columns(df, context="two-digit dataframe")
    out = df.copy()
    if "date" not in out.columns:
        raise ValueError("two-digit dataframe is missing date")
    out["date"] = out["date"].apply(_fmt_date)
    for col in FIELD_WIDTHS:
        out[col] = out[col].apply(lambda v, field=col: _fmt_two_digit(v, field))
    return out


def _write_df(ws, df: pd.DataFrame, *, freeze_row: int = 1) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    text_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = text_align

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, str):
                # Store lottery codes as text, never as numeric cells with only
                # cosmetic formatting; this survives workbook reopen/import.
                cell.number_format = "@"
            cell.alignment = text_align

    ws.freeze_panes = f"A{freeze_row + 1}"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, col in enumerate(df.columns, start=1):
        if df.empty:
            max_len = len(str(col))
        else:
            max_len = max(
                [
                    len(str(col)),
                    *[
                        len(str(v))
                        for v in df.iloc[:500, col_idx - 1].tolist()
                    ],
                ]
            )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(10, max_len + 2), 22
        )


def _verify_sheet_roundtrip(path: Path, sheet: str, df: pd.DataFrame) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False, read_only=False)
    ws = wb[sheet]
    headers = [cell.value for cell in ws[1]]
    expected_headers = [str(c) for c in df.columns]
    if headers != expected_headers:
        raise RuntimeError(
            f"Excel header mismatch in {path.name}/{sheet}: "
            f"{headers!r} != {expected_headers!r}"
        )
    if ws.max_row != len(df) + 1:
        raise RuntimeError(
            f"Excel row mismatch in {path.name}/{sheet}: "
            f"{ws.max_row - 1} != {len(df)}"
        )
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, expected in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx)
            if isinstance(expected, str):
                # OOXML/openpyxl serializes an empty string as a blank cell and
                # reads it back as None. This is acceptable only for genuinely
                # empty display cells (for example a loto head with no tails).
                if expected == "" and cell.value is None:
                    continue
                if not isinstance(cell.value, str) or cell.value != expected:
                    raise RuntimeError(
                        f"Excel text roundtrip mismatch {path.name}/{sheet} "
                        f"R{r_idx}C{c_idx}: {cell.value!r} != {expected!r}"
                    )
                if cell.number_format != "@":
                    raise RuntimeError(
                        f"Excel text format missing at {path.name}/{sheet} "
                        f"R{r_idx}C{c_idx}"
                    )


def _save_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for name, df in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        _write_df(ws, df)

    wb.save(path)
    for name, df in sheets.items():
        _verify_sheet_roundtrip(path, name[:31], df)


def _daily_loto_table(row: pd.Series) -> pd.DataFrame:
    values = [int(_fmt_prize(row[field], field)[-2:]) for field in FIELD_WIDTHS]
    rows: list[dict[str, str]] = []
    for head in range(10):
        tails = sorted(v % 10 for v in values if v // 10 == head)
        rows.append({"Đầu": str(head), "Đuôi": ", ".join(str(t) for t in tails)})
    return pd.DataFrame(rows)


def _daily_prize_table(row: pd.Series) -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    for label, fields in PRIZE_GROUPS:
        rows.append(
            {
                "Giải": label,
                "Kết quả": ", ".join(_fmt_prize(row[f], f) for f in fields),
            }
        )
    return pd.DataFrame(rows)


def _export_daily_files(
    raw_df: pd.DataFrame, out_dir: Path, *, latest_only: bool
) -> None:
    if raw_df.empty:
        return

    _require_prize_columns(raw_df, context="daily raw dataframe")
    view = raw_df.sort_values("date")
    if latest_only:
        view = view.tail(1)

    daily_dir = out_dir / "daily"
    if latest_only and daily_dir.exists():
        for old_file in daily_dir.glob("xsmb_*.xlsx"):
            old_file.unlink(missing_ok=True)
    for _, row in view.iterrows():
        d = _fmt_date(row["date"])
        _save_workbook(
            daily_dir / f"xsmb_{d}.xlsx",
            {
                "Ket qua": _daily_prize_table(row),
                "Lo to dau duoi": _daily_loto_table(row),
            },
        )


def _validate_prize_sheet(path: Path, *, two_digit: bool) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False, read_only=False)
    sheet = "TwoDigits" if two_digit else "Raw"
    ws = wb[sheet]
    header = [cell.value for cell in ws[1]]
    index = {str(name): i + 1 for i, name in enumerate(header)}
    missing = [field for field in FIELD_WIDTHS if field not in index]
    if missing:
        raise RuntimeError(f"{path.name} missing prize columns: {missing}")

    checked = 0
    for row_idx in range(2, ws.max_row + 1):
        for field, width in FIELD_WIDTHS.items():
            cell = ws.cell(row_idx, index[field])
            expected_width = 2 if two_digit else width
            value = cell.value
            if not isinstance(value, str):
                raise RuntimeError(
                    f"{path.name} {field} row {row_idx} is not stored as text"
                )
            if (
                len(value) != expected_width
                or not value.isascii()
                or not value.isdigit()
            ):
                raise RuntimeError(
                    f"{path.name} {field} row {row_idx} has invalid code {value!r}; "
                    f"expected exactly {expected_width} digits"
                )
            if cell.number_format != "@":
                raise RuntimeError(
                    f"{path.name} {field} row {row_idx} is not Excel text format"
                )
            checked += 1
    return checked


def _validate_daily_workbook(path: Path) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False, read_only=False)
    ws = wb["Ket qua"]
    labels = {label: fields for label, fields in PRIZE_GROUPS}
    seen: set[str] = set()
    values_checked = 0
    for row_idx in range(2, ws.max_row + 1):
        label = ws.cell(row_idx, 1).value
        raw = ws.cell(row_idx, 2).value
        if label not in labels:
            raise RuntimeError(f"{path.name} has unknown prize label {label!r}")
        if label in seen:
            raise RuntimeError(f"{path.name} duplicates prize label {label!r}")
        seen.add(label)
        if not isinstance(raw, str):
            raise RuntimeError(f"{path.name} {label} result is not text")
        parts = raw.split(", ") if raw else []
        fields = labels[label]
        if len(parts) != len(fields):
            raise RuntimeError(
                f"{path.name} {label} has {len(parts)} result(s), expected {len(fields)}"
            )
        for value, field in zip(parts, fields):
            width = FIELD_WIDTHS[field]
            if len(value) != width or not value.isascii() or not value.isdigit():
                raise RuntimeError(
                    f"{path.name} {label} contains invalid {value!r}; "
                    f"expected {width} digits"
                )
            values_checked += 1
    if seen != set(labels):
        raise RuntimeError(
            f"{path.name} is missing prize group(s): {set(labels) - seen}"
        )
    if values_checked != EXPECTED_PRIZE_VALUES:
        raise RuntimeError(
            f"{path.name} has {values_checked} prize values, expected 27"
        )
    return values_checked


def validate_excel_outputs(data_dir: Path) -> dict[str, int]:
    """Verify workbooks preserve every lottery code and all 27 prize values."""

    excel_dir = data_dir / "excel"
    raw_path = excel_dir / "xsmb.xlsx"
    two_path = excel_dir / "xsmb-2-digits.xlsx"
    if not raw_path.exists() or not two_path.exists():
        raise RuntimeError("required Excel workbooks are missing")

    counts = {
        "raw_prize_cells": _validate_prize_sheet(raw_path, two_digit=False),
        "two_digit_prize_cells": _validate_prize_sheet(two_path, two_digit=True),
    }
    daily_files = sorted((excel_dir / "daily").glob("xsmb_*.xlsx"))
    if not daily_files:
        raise RuntimeError("daily Excel workbook is missing")
    counts["daily_prize_values"] = _validate_daily_workbook(daily_files[-1])
    return counts


def export_excel_outputs(
    *,
    raw_df: pd.DataFrame,
    two_digit_df: pd.DataFrame,
    sparse_df: pd.DataFrame,
    data_dir: Path,
    latest_daily_only: bool = False,
) -> None:
    """Export Excel workbooks with exact-width lottery codes.

    Prize cells are stored as strings with Excel text format ``@``. The writer
    validates all 27 prize fields and immediately reopens every workbook to
    prove that leading zeroes and exact field widths survived serialization.
    """

    try:
        import openpyxl  # noqa: F401
    except Exception as exc:  # pragma: no cover
        logger.warning("openpyxl is not available; skip Excel export: %s", exc)
        return

    raw_text = _format_raw_df(raw_df)
    two_text = _format_two_digit_df(two_digit_df)

    excel_dir = data_dir / "excel"
    _save_workbook(excel_dir / "xsmb.xlsx", {"Raw": raw_text})
    _save_workbook(
        excel_dir / "xsmb-2-digits.xlsx", {"TwoDigits": two_text}
    )

    if not sparse_df.empty:
        sparse = sparse_df.copy()
        if "date" in sparse.columns:
            sparse["date"] = sparse["date"].apply(_fmt_date)
        _save_workbook(excel_dir / "xsmb-sparse.xlsx", {"Sparse": sparse})

    _export_daily_files(raw_df, excel_dir, latest_only=latest_daily_only)
    counts = validate_excel_outputs(data_dir)
    logger.info("Excel integrity verified: %s", counts)
