from __future__ import annotations

"""Validate committed Excel outputs against canonical XSMB data.

This validator checks *values*, not only Excel formatting. Every prize code in
``xsmb.xlsx`` and ``xsmb-2-digits.xlsx`` must exactly match ``data/xsmb.csv``
after canonical width formatting, and the latest daily workbook must contain all
8 prize groups / 27 prize values without losing leading zeroes.
"""

import argparse
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from excel_export import (
    EXPECTED_PRIZE_VALUES,
    FIELD_WIDTHS,
    PRIZE_GROUPS,
    _fmt_date,
    _fmt_prize,
    validate_excel_outputs,
)


def _sheet_table(path: Path, sheet: str) -> pd.DataFrame:
    wb = load_workbook(path, data_only=False, read_only=False)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError(f"empty Excel sheet: {path.name}/{sheet}")
    header = [str(x) if x is not None else "" for x in rows[0]]
    if len(header) != len(set(header)):
        raise RuntimeError(f"duplicate Excel headers: {path.name}/{sheet}")
    return pd.DataFrame(rows[1:], columns=header)


def _canonical_raw(csv_path: Path) -> pd.DataFrame:
    raw = pd.read_csv(csv_path)
    required = ["date", *FIELD_WIDTHS]
    missing = [c for c in required if c not in raw.columns]
    if missing:
        raise RuntimeError(f"canonical CSV missing fields: {missing}")
    out = raw[required].copy()
    out["date"] = out["date"].map(_fmt_date)
    for field in FIELD_WIDTHS:
        out[field] = out[field].map(lambda v, f=field: _fmt_prize(v, f))
    return out


def _canonical_two(raw: pd.DataFrame) -> pd.DataFrame:
    out = raw.copy()
    for field in FIELD_WIDTHS:
        out[field] = out[field].map(lambda v: str(v)[-2:])
    return out


def _assert_frame_exact(actual: pd.DataFrame, expected: pd.DataFrame, *, label: str) -> None:
    cols = list(expected.columns)
    if list(actual.columns) != cols:
        raise RuntimeError(f"{label} column mismatch: {list(actual.columns)!r} != {cols!r}")
    if len(actual) != len(expected):
        raise RuntimeError(f"{label} row mismatch: {len(actual)} != {len(expected)}")
    for col in cols:
        a = actual[col].fillna("").astype(str).tolist()
        e = expected[col].fillna("").astype(str).tolist()
        if a != e:
            for i, (av, ev) in enumerate(zip(a, e), start=2):
                if av != ev:
                    raise RuntimeError(
                        f"{label} value mismatch at row {i}, column {col}: {av!r} != {ev!r}"
                    )
            raise RuntimeError(f"{label} values differ in column {col}")


def _daily_expected(row: pd.Series) -> dict[str, list[str]]:
    return {
        label: [_fmt_prize(row[field], field) for field in fields]
        for label, fields in PRIZE_GROUPS
    }


def _validate_latest_daily(path: Path, latest: pd.Series) -> int:
    wb = load_workbook(path, data_only=False, read_only=False)
    ws = wb["Ket qua"]
    actual: dict[str, list[str]] = {}
    for r_idx in range(2, ws.max_row + 1):
        label = ws.cell(r_idx, 1).value
        value = ws.cell(r_idx, 2).value
        if not isinstance(label, str) or not isinstance(value, str):
            raise RuntimeError(f"daily workbook row {r_idx} is not textual")
        actual[label] = value.split(", ") if value else []

    expected = _daily_expected(latest)
    if actual != expected:
        raise RuntimeError(f"latest daily workbook differs from canonical draw: {actual!r} != {expected!r}")
    total = sum(len(v) for v in actual.values())
    if total != EXPECTED_PRIZE_VALUES:
        raise RuntimeError(f"daily workbook contains {total} prize values, expected 27")
    return total


def validate_against_canonical(*, data_dir: Path) -> dict[str, object]:
    csv_path = data_dir / "xsmb.csv"
    if not csv_path.exists():
        raise RuntimeError(f"canonical data missing: {csv_path}")

    structural = validate_excel_outputs(data_dir)
    expected_raw = _canonical_raw(csv_path)
    expected_two = _canonical_two(expected_raw)

    raw_path = data_dir / "excel" / "xsmb.xlsx"
    two_path = data_dir / "excel" / "xsmb-2-digits.xlsx"
    actual_raw = _sheet_table(raw_path, "Raw")
    actual_two = _sheet_table(two_path, "TwoDigits")
    _assert_frame_exact(actual_raw, expected_raw, label="xsmb.xlsx/Raw")
    _assert_frame_exact(actual_two, expected_two, label="xsmb-2-digits.xlsx/TwoDigits")

    latest = expected_raw.iloc[-1]
    latest_date = str(latest["date"])
    daily_path = data_dir / "excel" / "daily" / f"xsmb_{latest_date}.xlsx"
    if not daily_path.exists():
        raise RuntimeError(f"latest daily workbook missing: {daily_path}")
    daily_values = _validate_latest_daily(daily_path, latest)

    expected_cells = len(expected_raw) * EXPECTED_PRIZE_VALUES
    if structural["raw_prize_cells"] != expected_cells:
        raise RuntimeError(
            f"raw workbook checked {structural['raw_prize_cells']} prize cells, expected {expected_cells}"
        )
    if structural["two_digit_prize_cells"] != expected_cells:
        raise RuntimeError(
            "two-digit workbook prize-cell count differs from canonical row count"
        )

    return {
        "ok": True,
        "latest_date": latest_date,
        "canonical_rows": len(expected_raw),
        "prize_values_per_draw": EXPECTED_PRIZE_VALUES,
        "raw_prize_cells_checked": structural["raw_prize_cells"],
        "two_digit_prize_cells_checked": structural["two_digit_prize_cells"],
        "latest_daily_prize_values_checked": daily_values,
        "leading_zero_policy": "lottery codes stored as Excel text with exact widths 5/4/3/2",
    }


def main() -> None:
    ap = argparse.ArgumentParser(description="Validate Excel lottery data against canonical CSV.")
    ap.add_argument("--data-dir", default="data")
    ap.add_argument("--json-out", default="data/excel/integrity.json")
    args = ap.parse_args()
    report = validate_against_canonical(data_dir=Path(args.data_dir))
    out = Path(args.json_out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print("[OK] Excel integrity:", report)


if __name__ == "__main__":
    main()
